import logging
import re
import pandas as pd
import datetime
from pathlib import Path
from openpyxl import load_workbook

def logging_setup(module_name = 'default'):
    # Set up logging to console
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)

    # Set up logging to file
    file_handler = logging.FileHandler('log.txt') # overwrite existing file
    file_handler.setLevel(logging.INFO)

    # Create a logger and set the logging level
    logger = logging.getLogger(module_name)
    logger.setLevel(logging.DEBUG)

    # Add the console and file handlers to the logger
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    # create a formatter and add it to the handler
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(formatter)
    file_handler.setFormatter(formatter)
    return logger

# Log to excel
def save_pred(command, average, stddev, rand, config, logger):
    
    # Get data_config.txt from libfm folder
    folder_path = Path(f'{config["PATHS"]["libfm_path"]}')
    dataconfig = Path(folder_path, 'data_config.txt')
    # Check if data_config.txt exists
    if not dataconfig.is_file():
        n = 'Unknown'
        features = 'Unknown'
    else:
        with open(dataconfig, 'r') as f:
            dataconfig = f.read()
        # Check if either '-FULL DATA-' or 'Instances' is in dataconfig.txt
        if '-FULL DATA-' in dataconfig:
            n = 'FULL DATA'
        elif 'Instances' in dataconfig:
            # Get number behind 'Instances'
            n = dataconfig.split('Instances: ')[1].split('  ')[0].split('\n')[0]
        else:
            n = 'Unknown'

        # Extract feature names and values using regular expressions
        feature_dict = {}
        for match in re.finditer(r'^\s*(.+)\s+:\s+(\w+)', dataconfig, flags=re.MULTILINE):
            feature_name, feature_value = match.group(1), match.group(2)
            key = feature_name.strip().replace(' ', '') # remove whitespaces from the key
            value = feature_value.strip() # remove whitespaces from the value
            if key != 'Created' and key != 'Version' and key != 'Grouping':
                feature_dict[key] = value

        # Iterate over dict. if key has no _abbreviation, get value off key with abbreviation and add to list
        feature_string = ''
        for key, value in feature_dict.items():
            if 'Abbreviation' not in key and (value == 'true' or value == 'True'):
                if key + 'Abbreviation' in feature_dict.keys():
                    feature_string += feature_dict[key + 'Abbreviation']
                else:
                    logger.warning(f'No abbreviation for {key} found')
        if feature_string == '':
            feature_string = 'No features selected'

    today = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    df = pd.DataFrame({'date': [today], 'command': [command], 'Average': [average],\
                       'Standard Deviation': [stddev], 'random': [rand], 'Number Instances': [n], 'Features': [feature_string]})

    filename = 'Runs.xlsx'

    # Check if the file exists
    try:
        with open(filename) as f:
            pass
        # Load the workbook and sheet
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl') 
        writer.book = book
        
        # Append the DataFrame to the existing sheet
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        
        # Save the changes
        writer.save()
        writer.close()

    except FileNotFoundError:
        # Create a new workbook and sheet
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        writer.close()
    return